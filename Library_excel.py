import random
import pandas as pd
from flask import Flask, request, send_file, jsonify
from datetime import datetime
import sqlite3
import io
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl import Workbook
import os
import traceback

# --- 1. APPLICATION SETUP ---
app = Flask(__name__)
DATABASE = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'staff.db')

# --- 2. CONFIGURATION DATA ---
ROLE_PRIORITY = {"Duty Manager": 1, "Scale 3": 2, "Volunteer": 3}

TASK_CONFIG = {
    "SM":         {"roles": ["Scale 3"], "mandatory": 1, "full_name": "SM"},
    "R":          {"roles": ["Scale 3"], "mandatory": 1, "full_name": "R"},
    "C":          {"roles": ["Scale 3"], "mandatory": 0, "full_name": "C (C/AV)"},
    "C+":         {"roles": ["Scale 3"], "mandatory": 0, "full_name": "C+ (A/T)"},
    "1st":        {"roles": ["Scale 3", "Volunteer"], "mandatory": 0, "full_name": "1st"},
    "Res":        {"roles": ["Scale 3", "Volunteer"], "mandatory": 0, "full_name": "Res"},
    "Set Up":     {"roles": ["Duty Manager"], "mandatory": 0, "full_name": "Set Up"},
    "T":          {"roles": ["Scale 3", "Duty Manager", "Volunteer"], "mandatory": 0, "full_name": "T"},
}
MANDATORY_C_COVERAGE = 2

GREEN_FILL = PatternFill(start_color="92D050",
                         end_color="92D050", fill_type="solid")

BLACKOUT_FILL = PatternFill(
    start_color="000000", end_color="000000", fill_type="solid")

# First staff row in Excel
DATA_START_ROW_EXCEL = 6


# --- 3. DATABASE SETUP AND STAFF MANAGEMENT ---


def init_db():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS staff (
            id INTEGER PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            role TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()


@app.route('/staff', methods=['GET', 'POST'])
def manage_staff():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()

    if request.method == 'POST':
        data = request.json
        name = data.get('name')
        role = data.get('role')
        if not name or not role:
            conn.close()
            return jsonify({"error": "Name and Role are required."}), 400

        try:
            c.execute("INSERT INTO staff (name, role) VALUES (?, ?)",
                      (name, role))
            conn.commit()
            conn.close()
            return jsonify({"message": f"Staff member {name} added as {role}."}), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({"error": f"Staff member {name} already exists."}), 409

    staff_list = c.execute("SELECT name, role FROM staff").fetchall()
    conn.close()

    return jsonify([{"name": s[0], "role": s[1]} for s in staff_list])


# --- 4. CORE SCHEDULING LOGIC ---


def auto_assign_tea_slots(staff_data):
    """
    Automatically assign tea slots for everyone on shift 13:00–14:00.
    Spread across 13:00, 13:15, 13:30, 13:45.
    """
    tea_times = ["13:00", "13:15", "13:30", "13:45"]
    idx = 0

    for staff in staff_data:
        if staff.get("status", "Available") != "Available":
            continue

        start = staff.get("start_hour", 0)
        end = staff.get("end_hour", 0)

        if start <= 13 < end:
            staff["tea_slot"] = tea_times[idx % len(tea_times)]
            idx += 1


def build_shift_label(staff):
    """
    Build 'Shift' text like the sheet: e.g. '11.30-4', '12-4', '2-4'.
    """
    role = staff.get("role", "")
    start = staff.get("start_hour", 12)
    end = staff.get("end_hour", 16)

    if role == "Duty Manager" and start <= 11:
        start_str = "11.30"
    else:
        start_str = str(start)

    end_str = str(end - 12) if end > 12 else str(end)
    return f"{start_str}-{end_str}"


def generate_schedule_data(staff_data, date_str):
    # Assign teas automatically
    auto_assign_tea_slots(staff_data)

    # Sort staff by availability, role, name
    def sort_key(s):
        availability_priority = 1 if s.get(
            "status", "Available") == "Available" else 2
        role_p = ROLE_PRIORITY.get(s.get("role", ""), 99)
        return (availability_priority, role_p, s.get("name", ""))

    staff_data.sort(key=sort_key)
    pivot_schedule = {s["name"]: {} for s in staff_data if "name" in s}

    # Track who has already had SM (stock movement) today
    sm_assigned_staff = set()

    # --- 2. Fixed Assignments (Set Up, Tea) ---
    setup_slot_key = "11:30"

    for staff in staff_data:
        if staff.get("status", "Available") != "Available":
            continue

        start = staff.get("start_hour", 0)
        end = staff.get("end_hour", 0)
        role = staff.get("role", "")
        name = staff.get("name", "")

        # Set Up (11:30–12) for DMs
        if role == "Duty Manager" and (start <= 11 < end):
            pivot_schedule[name][setup_slot_key] = "Set Up"
        # NOTE: we no longer write T into pivot_schedule here;
        # we only track tea via staff["tea_slot"] and handle it in the display layer.

    # --- 3. Per-hour tasks ---
    assignment_hour_keys = ["12:00", "13:00", "14:00", "15:00"]

    for time_str in assignment_hour_keys:
        hour = int(time_str.split(':')[0])

        available_staff_for_hour = []
        for s in staff_data:
            if s.get("status", "Available") != "Available":
                continue

            start = s.get("start_hour", 0)
            end = s.get("end_hour", 0)

            is_on_shift = start <= hour < end

            # IMPORTANT: do NOT exclude staff on tea here.
            # They still have a base task for the hour; tea is just 15 minutes.
            if is_on_shift:
                available_staff_for_hour.append(s)

        random.shuffle(available_staff_for_hour)
        tasks_assigned_in_hour = []

        # Mandatory cover: SM (one per hour, but each person at most once per day),
        # R, C, C+
        mandatory_tasks_to_assign = []

        # SM mandatory every hour, subject to per-person limit
        if TASK_CONFIG["SM"]["mandatory"] > 0:
            mandatory_tasks_to_assign.append("SM")

        # R mandatory every hour
        if TASK_CONFIG["R"]["mandatory"] > 0:
            mandatory_tasks_to_assign.append("R")

        # C and C+ for coverage
        mandatory_tasks_to_assign.extend(["C", "C+"][:MANDATORY_C_COVERAGE])

        # Only Scale 3s get these mandatory tasks
        scale3_staff = [
            s for s in available_staff_for_hour if s.get("role") == "Scale 3"
        ]

        for task in mandatory_tasks_to_assign:
            if task == "SM":
                # Only consider Scale 3s who haven't had SM yet today
                candidates = [
                    s for s in scale3_staff if s["name"] not in sm_assigned_staff
                ]
                if not candidates:
                    # No-one left who hasn't had SM; skip this SM this hour
                    continue
                staff_to_assign = candidates[0]
                name = staff_to_assign["name"]
                sm_assigned_staff.add(name)
                # Remove from scale3_staff so they don't also get R/C in this hour
                if staff_to_assign in scale3_staff:
                    scale3_staff.remove(staff_to_assign)
            else:
                if not scale3_staff:
                    continue
                staff_to_assign = scale3_staff.pop(0)
                name = staff_to_assign["name"]

            pivot_schedule[name][time_str] = task
            tasks_assigned_in_hour.append(name)

        # Random tasks: 1st, Res
        remaining_staff = [
            s for s in available_staff_for_hour if s["name"] not in tasks_assigned_in_hour
        ]
        random_tasks = [
            t for t in TASK_CONFIG
            if TASK_CONFIG[t]["mandatory"] == 0 and t not in ["Set Up", "T", "C", "C+", "R", "SM"]
        ]

        for staff in remaining_staff:
            role = staff.get("role", "")
            name = staff.get("name", "")
            assignable_tasks = [
                t for t in random_tasks if role in TASK_CONFIG[t]["roles"]]
            if assignable_tasks:
                assigned_task = random.choice(assignable_tasks)
                pivot_schedule[name][time_str] = assigned_task
                tasks_assigned_in_hour.append(name)

    # --- 4. Build DataFrame for the table area ---

    display_time_headers = [
        "11.30-12",
        "12-1",
        "00",
        "15",
        "30",
        "45",
        "2-3",
        "3-4",
    ]

    internal_to_display_map = {
        "11:30": "11.30-12",
        "12:00": "12-1",
        "13:00": "1-2",      # handle 1–2 specially
        "14:00": "2-3",
        "15:00": "3-4",
    }

    pivot_df_data = []
    task_code_map = {k: v["full_name"] for k, v in TASK_CONFIG.items()}

    for staff in staff_data:
        name = staff.get("name", "")
        role = staff.get("role", "")
        shift_label = build_shift_label(staff)

        row = {"Staff Name": name, "Shift": shift_label}
        for header in display_time_headers:
            row[header] = ""
        row["Comments"] = ""

        status = staff.get("status", "Available")
        status_detail = staff.get("status_detail", "")

        # Non-available staff behaviour
        if status != "Available":
            if status == "Annual Leave":
                for header in display_time_headers:
                    row[header] = "A/L"
                row["Comments"] = "A/L"
            else:
                # Sick or Other Library:
                # leave time cells + Comments blank; blackout + merge later
                pass

        else:
            # Available staff: normal scheduling
            for internal_key, display_key in internal_to_display_map.items():
                # 11:30 Set Up (DM only)
                if internal_key == "11:30":
                    task_code = pivot_schedule.get(
                        name, {}).get(internal_key, "")
                    if task_code:
                        display_task = task_code_map.get(task_code, task_code)
                        row["11.30-12"] = display_task
                    continue

                # 13:00 – split into 4 mini slots (00/15/30/45)
                if internal_key == "13:00":
                    base_task_code = pivot_schedule.get(
                        name, {}).get(internal_key, "")
                    tea_slot = staff.get("tea_slot")

                    for minute, col_name in [("00", "00"), ("15", "15"),
                                             ("30", "30"), ("45", "45")]:
                        # If this is their tea minute, put T
                        if isinstance(tea_slot, str) and tea_slot == f"13:{minute}":
                            row[col_name] = task_code_map["T"]
                        else:
                            # Otherwise, use base task if there is one
                            if base_task_code:
                                display_task = task_code_map.get(
                                    base_task_code, base_task_code)
                                # DMs shouldn't show "normal" tasks
                                if not (role == "Duty Manager" and base_task_code not in ["Set Up", "T"]):
                                    row[col_name] = display_task
                    continue

                # Normal hourly blocks: 12–1, 2–3, 3–4
                task_code = pivot_schedule.get(name, {}).get(internal_key, "")
                if task_code:
                    display_task = task_code_map.get(task_code, task_code)
                    if role == "Duty Manager" and task_code not in ["Set Up", "T"]:
                        continue

                    if internal_key == "12:00":
                        row["12-1"] = display_task
                    elif internal_key == "14:00":
                        row["2-3"] = display_task
                    elif internal_key == "15:00":
                        row["3-4"] = display_task

        pivot_df_data.append(row)

    df = pd.DataFrame(pivot_df_data)

    final_columns_names = [
        "Staff Name",
        "Shift",
        "11.30-12",
        "12-1",
        "00",
        "15",
        "30",
        "45",
        "2-3",
        "3-4",
        "Comments",
    ]
    df = df[final_columns_names]

    return df


# --- 5. EXCEL HEADER ROWS, MERGING, STYLING, BLACKOUTS, LIBRARIES/SICK ---


def add_template_header_rows(worksheet, date_obj, duty_manager_names):
    """
    Header layout (columns A–K):

    Row 1: merged A1:K1 – title
    Row 2: merged A2:K2 – blank for notes
    Row 3: A3 = date, B3:C3 blank, D3–H3 "Duty Manager(s)", I3–K3 DM names
    Row 4: main headers: Name, Shift, 11.30-12, 12-1, 1-2 (merged E4–H4), 2-3, 3-4, Comments
    Row 5: 00, 15, 30, 45 under 1–2 (E5–H5)
    """

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 10
    worksheet.column_dimensions["D"].width = 10
    for col in ["E", "F", "G", "H"]:
        worksheet.column_dimensions[col].width = 3.2
    worksheet.column_dimensions["I"].width = 10
    worksheet.column_dimensions["J"].width = 10
    worksheet.column_dimensions["K"].width = 31.5

    SIZE = 10

    # Row 1: title
    worksheet.merge_cells("A1:K1")
    cell = worksheet["A1"]
    cell.value = "Canada Water Library Sunday week 3"
    cell.font = Font(name="Arial", bold=True)

    # Row 2: notes
    worksheet.merge_cells("A2:K2")
    worksheet["A2"].font = Font(name="Arial")

    # Row 3: date + duty manager bands
    worksheet.merge_cells("B3:C3")

    short_date = date_obj.strftime("%d/%m/%y")
    worksheet["A3"] = short_date
    worksheet["A3"].font = Font(name="Arial", size=SIZE)

    # Only assign B3 (merged B3:C3)
    worksheet["B3"] = ""

    worksheet.merge_cells("D3:H3")
    worksheet["D3"] = "Duty Manager(s)"
    worksheet["D3"].font = Font(name="Arial", bold=True, size=SIZE)

    worksheet.merge_cells("I3:K3")
    worksheet["I3"] = duty_manager_names
    worksheet["I3"].font = Font(name="Arial", bold=True, size=SIZE)
    worksheet["I3"].alignment = Alignment(
        horizontal="center", vertical="center")

    # Row 4 headers
    headers_row4 = [
        ("A4", "Name"),
        ("B4", "Shift"),
        ("C4", "11.30-12"),
        ("D4", "12-1"),
        ("I4", "2-3"),
        ("J4", "3-4"),
        ("K4", "Comments"),
    ]
    for cell_ref, text in headers_row4:
        c = worksheet[cell_ref]
        c.value = text
        c.font = Font(name="Arial", bold=True, size=SIZE)
        c.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("E4:H4")
    worksheet["E4"] = "1-2"
    worksheet["E4"].font = Font(name="Arial", bold=True, size=SIZE)
    worksheet["E4"].alignment = Alignment(
        horizontal="center", vertical="center")

    # Row 5 mini headers
    sub_headers = [
        ("E5", "00"),
        ("F5", "15"),
        ("G5", "30"),
        ("H5", "45"),
    ]
    for cell_ref, text in sub_headers:
        cell = worksheet[cell_ref]
        cell.value = text
        cell.font = Font(name="Arial", bold=True, size=SIZE)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Centre entire row 4
    alignment_center = Alignment(horizontal="center", vertical="center")
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        cell = worksheet[f"{col_letter}4"]
        cell.alignment = alignment_center


def apply_excel_styling(writer, df, staff_data):
    workbook = writer.book
    worksheet = writer.sheets['Timesheet']

    # Global Arial font
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial")

    sm_display_name = TASK_CONFIG["SM"]["full_name"]
    staff_by_name = {s["name"]: s for s in staff_data if "name" in s}

    time_columns = {
        "11.30-12": 11.5,
        "12-1": 12,
        "00": 13,
        "15": 13,
        "30": 13,
        "45": 13,
        "2-3": 14,
        "3-4": 15,
    }
    time_col_order = ["11.30-12", "12-1", "00", "15", "30", "45", "2-3", "3-4"]

    col_indices = {col: idx for idx, col in enumerate(df.columns)}
    first_time_col_idx = col_indices["11.30-12"]
    comments_col_idx = col_indices["Comments"]

    # 1. Bold names/tasks, green SM
    for r_idx, row in df.iterrows():
        excel_row = DATA_START_ROW_EXCEL + r_idx

        name_cell = worksheet.cell(row=excel_row, column=1)
        name_cell.font = Font(name="Arial", bold=True)

        for c_idx in range(first_time_col_idx, comments_col_idx + 1):
            cell_value = row.iloc[c_idx]
            excel_col = c_idx + 1
            cell = worksheet.cell(row=excel_row, column=excel_col)

            if cell_value:
                cell.font = Font(name="Arial", bold=True)

            if cell_value == sm_display_name:
                cell.fill = GREEN_FILL

    # 2. Blackout outside shift
    for r_idx, row in df.iterrows():
        name = row["Staff Name"]
        staff_info = staff_by_name.get(name)
        if not staff_info:
            continue

        start_hour = staff_info.get("start_hour", 0)
        end_hour = staff_info.get("end_hour", 24)
        excel_row = DATA_START_ROW_EXCEL + r_idx

        for col_name, hour_value in time_columns.items():
            c_idx = col_indices[col_name]
            cell = worksheet.cell(row=excel_row, column=c_idx + 1)
            if hour_value < start_hour or hour_value >= end_hour:
                cell.fill = BLACKOUT_FILL

    # 3. Merge & colour Sick/other library rows
    for r_idx, row in df.iterrows():
        name = row["Staff Name"]
        staff_info = staff_by_name.get(name)
        if not staff_info:
            continue

        status = staff_info.get("status", "Available")
        status_detail = staff_info.get("status_detail", "")

        is_special = False
        label = ""

        if status == "Sick":
            is_special = True
            label = "SICK"
        elif status not in ["Available", "Annual Leave"] and status_detail:
            is_special = True
            label = status_detail

        if not is_special:
            continue

        start_hour = staff_info.get("start_hour", 12)
        end_hour = staff_info.get("end_hour", 16)
        excel_row = DATA_START_ROW_EXCEL + r_idx

        cols_to_merge = []
        for col_name in time_col_order:
            hour_value = time_columns[col_name]
            if start_hour <= hour_value < end_hour:
                cols_to_merge.append(col_name)

        if not cols_to_merge:
            continue

        first_col_name = cols_to_merge[0]
        last_col_name = cols_to_merge[-1]

        first_col_idx = col_indices[first_col_name] + 1
        last_col_idx = col_indices[last_col_name] + 1

        worksheet.merge_cells(
            start_row=excel_row,
            start_column=first_col_idx,
            end_row=excel_row,
            end_column=last_col_idx
        )

        rand_color = "{:02X}{:02X}{:02X}".format(
            random.randint(80, 240),
            random.randint(80, 240),
            random.randint(80, 240),
        )
        fill = PatternFill(start_color=rand_color,
                           end_color=rand_color, fill_type="solid")

        merged_cell = worksheet.cell(row=excel_row, column=first_col_idx)
        merged_cell.value = label
        merged_cell.fill = fill
        merged_cell.font = Font(name="Arial", bold=True)
        merged_cell.alignment = Alignment(
            horizontal="center", vertical="center")

    # 4. Borders A1:K<last staff row>
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    max_row = DATA_START_ROW_EXCEL + len(df) - 1
    for r in range(1, max_row + 1):
        for c in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=r, column=c)
            cell.border = thin_border

    # 5. Center-align D–J for staff rows
    alignment_center = Alignment(horizontal="center", vertical="center")
    for r_idx in range(DATA_START_ROW_EXCEL, DATA_START_ROW_EXCEL + len(df)):
        for col_letter in ["D", "E", "F", "G", "H", "I", "J"]:
            cell = worksheet[f"{col_letter}{r_idx}"]
            cell.alignment = alignment_center

    return workbook


def parse_date_from_payload(raw_date: str | None) -> datetime:
    if not raw_date:
        return datetime.now()
    try:
        return datetime.strptime(raw_date, "%Y-%m-%d")
    except ValueError:
        pass
    cleaned = raw_date.replace("Z", "").split("T")[0]
    return datetime.strptime(cleaned, "%Y-%m-%d")


@app.route('/generate-timesheet', methods=['POST'])
def generate_timesheet():
    data = request.json or {}
    staff_data = data.get('schedule', [])
    raw_date = data.get('date')

    if not staff_data:
        return jsonify({"error": "No staff data provided for scheduling."}), 400

    try:
        df = generate_schedule_data(staff_data, raw_date)

        duty_managers = [s["name"]
                         for s in staff_data if s.get("role") == "Duty Manager"]
        duty_manager_names = " & ".join(duty_managers)

        date_obj = parse_date_from_payload(raw_date)
        formatted_date = date_obj.strftime("%A, %d %B %Y")
        filename = f"Timesheet_{formatted_date}.xlsx"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(
                writer,
                sheet_name='Timesheet',
                index=False,
                header=False,
                startrow=DATA_START_ROW_EXCEL - 1
            )

            worksheet = writer.sheets['Timesheet']
            add_template_header_rows(worksheet, date_obj, duty_manager_names)
            apply_excel_styling(writer, df, staff_data)

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Scheduling Error: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate timesheet. Error: {str(e)}"}), 500


if __name__ == '__main__':
    init_db()
    print("\n---------------------------------------------------------")
    print("  Flask App Running. Use /staff to manage staff names.")
    print("  Use /generate-timesheet (POST) to create the schedule.")
    print("---------------------------------------------------------")
    app.run(debug=True)
